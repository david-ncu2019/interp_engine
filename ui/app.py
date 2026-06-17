"""
app.py  –  Main Tkinter application window.

4-tab Notebook:
  Tab 1 – Data        : file picker, column selection, output settings
  Tab 2 – Method      : engine selector (Kriging / GP)
  Tab 3 – Model       : variogram / kernel fitting panel (switches per method)
  Tab 4 – Run & Results : subprocess run, live log, summary metrics
"""

import os
import queue
import subprocess
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from typing import Optional

# Add project root to path so imports work when launched from anywhere
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from ui.engine_runner import AutoOptimizeRunner, EngineRunner
from ui.variogram_panel import KrigingPanel, GPPanel


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _read_columns(filepath: str) -> list[str]:
    """Return column header names from a CSV, XLSX, or SHP file."""
    ext = Path(filepath).suffix.lower()
    try:
        if ext in (".csv", ".txt"):
            import csv
            with open(filepath, newline="", encoding="utf-8-sig") as f:
                return next(csv.reader(f))
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            return [cell.value for cell in next(ws.iter_rows(max_row=1)) if cell.value is not None]
        elif ext == ".shp":
            import geopandas as gpd
            gdf = gpd.read_file(filepath)
            return list(gdf.columns)
    except Exception as exc:
        messagebox.showerror("Column read error", str(exc))
    return []


def _load_xyz(filepath: str, col_x: str, col_y: str, col_z: str):
    """Load X, Y, Z arrays from file for empirical variogram computation."""
    import numpy as np
    ext = Path(filepath).suffix.lower()
    try:
        if ext in (".csv", ".txt"):
            import csv
            rows = []
            with open(filepath, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    rows.append((float(r[col_x]), float(r[col_y]), float(r[col_z])))
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            ix, iy, iz = headers.index(col_x), headers.index(col_y), headers.index(col_z)
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[ix] is not None:
                    rows.append((float(row[ix]), float(row[iy]), float(row[iz])))
        elif ext == ".shp":
            import geopandas as gpd
            gdf = gpd.read_file(filepath)
            rows = list(zip(gdf[col_x], gdf[col_y], gdf[col_z]))
        else:
            return None, None, None
        arr = np.array(rows, dtype=float)
        return arr[:, :2], arr[:, 2], len(arr)
    except Exception:
        return None, None, None


# ─────────────────────────────────────────────────────────────────────────────
# Main application
# ─────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Interpolation Engine")
        self.geometry("1280x800")
        self.minsize(900, 600)

        # ── Shared state dict ─────────────────────────────────────────────────
        self.state: dict = {
            "input_filepath":     "",
            "col_x":              "",
            "col_y":              "",
            "col_value":          "",
            "ground_truth_filepath": "",
            "output_dir":         str(PROJECT_ROOT / "output"),
            "export_formats":     ["nc"],
            "resolution_m":       50.0,
            "convex_hull_buffer": 10.0,
            "detrend_auto":       True,
            "detrend_enabled":    True,
            "detrend_order":      1,
            "nst_enabled":        None,
            "min_separation":     None,
            "engine_mode":        "kriging",
            "kriging_n_trials":   300,
            "kriging_n_splits":   3,
            "kriging_max_anisotropy": 10.0,
            "gp_n_trials":        300,
            "gp_max_anisotropy":  15.0,
            "gp_angle_min":       0.0,
            "gp_angle_max":       180.0,
            "gp_random_state":    42,
            "save_diagnostics":   True,
            "netcdf_z_dim_name":  "Depth",
            "kriging_preset":     None,
            "gp_preset":          None,
        }

        # ── Runtime ────────────────────────────────────────────────────────────
        self._log_queue: queue.Queue = queue.Queue()
        self._runner: Optional[EngineRunner] = None
        self._opt_runner: Optional[AutoOptimizeRunner] = None

        # ── Build UI ──────────────────────────────────────────────────────────
        self._build_menu()
        self._notebook = ttk.Notebook(self)
        self._notebook.pack(fill="both", expand=True, padx=8, pady=8)

        self._tab_data    = self._build_tab_data()
        self._tab_method  = self._build_tab_method()
        self._tab_model   = self._build_tab_model()
        self._tab_run     = self._build_tab_run()

        self._notebook.add(self._tab_data,   text=" 1. Data ")
        self._notebook.add(self._tab_method, text=" 2. Method ")
        self._notebook.add(self._tab_model,  text=" 3. Model ")
        self._notebook.add(self._tab_run,    text=" 4. Run & Results ")

        # Bind panel event from variogram panel buttons
        self.bind_all("<<AutoOptimize>>", self._on_auto_optimize_requested)

    # ─────────────────────────────────────────────────────────────────────────
    # Menu bar
    # ─────────────────────────────────────────────────────────────────────────

    def _build_menu(self):
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open Input File…", command=self._browse_input)
        file_menu.add_command(label="Set Output Folder…", command=self._browse_output)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.destroy)
        menubar.add_cascade(label="File", menu=file_menu)
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About", command=lambda: messagebox.showinfo(
            "About", "Interpolation Engine UI\nGeostatistics · Kriging · Gaussian Process"))
        menubar.add_cascade(label="Help", menu=help_menu)
        self.config(menu=menubar)

    # ─────────────────────────────────────────────────────────────────────────
    # Tab 1 — Data
    # ─────────────────────────────────────────────────────────────────────────

    def _build_tab_data(self) -> ttk.Frame:
        tab = ttk.Frame(self._notebook, padding=12)

        # ── Input file row ────────────────────────────────────────────────────
        ttk.Label(tab, text="Input file:").grid(row=0, column=0, sticky="w")
        self._input_path_var = tk.StringVar()
        ttk.Entry(tab, textvariable=self._input_path_var, width=52).grid(
            row=0, column=1, sticky="ew", padx=4)
        ttk.Button(tab, text="Browse…", command=self._browse_input).grid(
            row=0, column=2, padx=4)

        # ── Column dropdowns ──────────────────────────────────────────────────
        ttk.Label(tab, text="X column:").grid(row=1, column=0, sticky="w", pady=(8, 2))
        self._col_x_var = tk.StringVar()
        self._col_x_cb  = ttk.Combobox(tab, textvariable=self._col_x_var,
                                         state="readonly", width=20)
        self._col_x_cb.grid(row=1, column=1, sticky="w", padx=4)
        self._col_x_cb.bind("<<ComboboxSelected>>",
                             lambda _: self._on_column_changed())

        ttk.Label(tab, text="Y column:").grid(row=2, column=0, sticky="w", pady=2)
        self._col_y_var = tk.StringVar()
        self._col_y_cb  = ttk.Combobox(tab, textvariable=self._col_y_var,
                                         state="readonly", width=20)
        self._col_y_cb.grid(row=2, column=1, sticky="w", padx=4)
        self._col_y_cb.bind("<<ComboboxSelected>>",
                             lambda _: self._on_column_changed())

        ttk.Label(tab, text="Value column:").grid(row=3, column=0, sticky="w", pady=2)
        self._col_val_var = tk.StringVar()
        self._col_val_cb  = ttk.Combobox(tab, textvariable=self._col_val_var,
                                           state="readonly", width=20)
        self._col_val_cb.grid(row=3, column=1, sticky="w", padx=4)
        self._col_val_cb.bind("<<ComboboxSelected>>",
                              lambda _: self._on_column_changed())

        # ── Ground truth (optional) ───────────────────────────────────────────
        ttk.Label(tab, text="Ground truth (optional):").grid(
            row=4, column=0, sticky="w", pady=(12, 2))
        self._gt_path_var = tk.StringVar()
        ttk.Entry(tab, textvariable=self._gt_path_var, width=52).grid(
            row=4, column=1, sticky="ew", padx=4)
        ttk.Button(tab, text="Browse…",
                   command=lambda: self._browse_any(self._gt_path_var,
                       "ground_truth_filepath")).grid(row=4, column=2, padx=4)

        # ── Output ────────────────────────────────────────────────────────────
        ttk.Label(tab, text="Output folder:").grid(
            row=5, column=0, sticky="w", pady=(8, 2))
        self._output_dir_var = tk.StringVar(value=self.state["output_dir"])
        ttk.Entry(tab, textvariable=self._output_dir_var, width=52).grid(
            row=5, column=1, sticky="ew", padx=4)
        ttk.Button(tab, text="Browse…", command=self._browse_output).grid(
            row=5, column=2, padx=4)
        self._output_dir_var.trace_add("write",
            lambda *_: self.state.update({"output_dir": self._output_dir_var.get()}))

        # ── Export formats ────────────────────────────────────────────────────
        ttk.Label(tab, text="Export formats:").grid(
            row=6, column=0, sticky="w", pady=(8, 0))
        fmt_frame = ttk.Frame(tab)
        fmt_frame.grid(row=6, column=1, sticky="w", pady=(8, 0))
        self._fmt_vars = {}
        for fmt, label in [("nc", "NetCDF"), ("tif", "GeoTIFF"), ("csv", "CSV")]:
            v = tk.BooleanVar(value=(fmt == "nc"))
            self._fmt_vars[fmt] = v
            ttk.Checkbutton(fmt_frame, text=label, variable=v,
                            command=self._sync_formats).pack(side="left", padx=6)

        # ── Grid resolution ───────────────────────────────────────────────────
        ttk.Label(tab, text="Grid resolution (m):").grid(
            row=7, column=0, sticky="w", pady=(8, 0))
        self._res_var = tk.StringVar(value="50")
        res_entry = ttk.Entry(tab, textvariable=self._res_var, width=10)
        res_entry.grid(row=7, column=1, sticky="w", padx=4, pady=(8, 0))
        self._res_var.trace_add("write", self._sync_resolution)

        # ── Advanced (collapsible) ────────────────────────────────────────────
        adv_toggle = ttk.Label(tab, text="▸ Advanced settings",
                               cursor="hand2", foreground="#0063cc")
        adv_toggle.grid(row=8, column=0, columnspan=3, sticky="w", pady=(12, 0))

        self._adv_frame = ttk.LabelFrame(tab, text="Advanced", padding=8)
        # Initially hidden
        adv_toggle.bind("<Button-1>", self._toggle_advanced)

        self._build_advanced(self._adv_frame)
        tab.columnconfigure(1, weight=1)
        return tab

    def _build_advanced(self, frame: ttk.Frame):
        ttk.Label(frame, text="Detrend:").grid(row=0, column=0, sticky="w")
        self._detrend_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame, text="Enabled", variable=self._detrend_var,
                        command=lambda: self.state.update(
                            {"detrend_enabled": self._detrend_var.get()})).grid(
            row=0, column=1, sticky="w")

        ttk.Label(frame, text="Detrend order:").grid(row=1, column=0, sticky="w", pady=2)
        self._detrend_order_var = tk.StringVar(value="1")
        ttk.Combobox(frame, textvariable=self._detrend_order_var,
                     values=["1", "2"], state="readonly", width=4).grid(
            row=1, column=1, sticky="w")
        self._detrend_order_var.trace_add("write",
            lambda *_: self.state.update({"detrend_order": int(self._detrend_order_var.get())}))

        ttk.Label(frame, text="NST enabled:").grid(row=2, column=0, sticky="w", pady=2)
        self._nst_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame, variable=self._nst_var,
                        command=lambda: self.state.update(
                            {"nst_enabled": self._nst_var.get() or None})).grid(
            row=2, column=1, sticky="w")

        ttk.Label(frame, text="Convex hull buffer (%):").grid(row=3, column=0, sticky="w", pady=2)
        self._hull_var = tk.StringVar(value="10")
        ttk.Entry(frame, textvariable=self._hull_var, width=6).grid(
            row=3, column=1, sticky="w")
        self._hull_var.trace_add("write", lambda *_: self._sync_hull())

        ttk.Label(frame, text="Min separation:").grid(row=4, column=0, sticky="w", pady=2)
        self._minsep_var = tk.StringVar(value="")
        ttk.Entry(frame, textvariable=self._minsep_var, width=8).grid(
            row=4, column=1, sticky="w")
        ttk.Label(frame, text="(blank = auto)", foreground="gray").grid(
            row=4, column=2, sticky="w", padx=4)
        self._minsep_var.trace_add("write", lambda *_: self._sync_minsep())

    def _toggle_advanced(self, event=None):
        label = event.widget
        if self._adv_frame.winfo_ismapped():
            self._adv_frame.grid_remove()
            label.config(text="▸ Advanced settings")
        else:
            self._adv_frame.grid(row=9, column=0, columnspan=3, sticky="ew", pady=(4, 0))
            label.config(text="▾ Advanced settings")

    # ─────────────────────────────────────────────────────────────────────────
    # Tab 2 — Method
    # ─────────────────────────────────────────────────────────────────────────

    def _build_tab_method(self) -> ttk.Frame:
        tab = ttk.Frame(self._notebook, padding=20)

        ttk.Label(tab, text="Choose interpolation engine:",
                  font=("TkDefaultFont", 11, "bold")).pack(anchor="w", pady=(0, 16))

        self._method_var = tk.StringVar(value="kriging")

        kriging_frame = ttk.LabelFrame(tab, text="Ordinary Kriging", padding=10)
        kriging_frame.pack(fill="x", pady=4)
        ttk.Radiobutton(kriging_frame, text="Use Kriging",
                        variable=self._method_var, value="kriging",
                        command=self._on_method_changed).pack(anchor="w")
        ttk.Label(kriging_frame,
                  text=("Classical geostatistical method. Fits a parametric variogram model\n"
                        "(spherical, exponential, Gaussian, …) then uses it to weight neighbors.\n"
                        "Best when you have strong spatial correlation and want an interpretable model."),
                  justify="left", foreground="#444").pack(anchor="w", padx=20, pady=(4, 0))

        gp_frame = ttk.LabelFrame(tab, text="Gaussian Process (GPR)", padding=10)
        gp_frame.pack(fill="x", pady=4)
        ttk.Radiobutton(gp_frame, text="Use Gaussian Process",
                        variable=self._method_var, value="gp",
                        command=self._on_method_changed).pack(anchor="w")
        ttk.Label(gp_frame,
                  text=("Machine-learning kernel regression with rigorous uncertainty estimates.\n"
                        "Optimizes anisotropy angle and length scales via marginal likelihood.\n"
                        "Best for complex, non-stationary fields or when comparing to ML baselines."),
                  justify="left", foreground="#444").pack(anchor="w", padx=20, pady=(4, 0))

        # ── Advanced hyperparameters (collapsed) ──────────────────────────────
        adv_label = ttk.Label(tab, text="▸ Optimization settings",
                               cursor="hand2", foreground="#0063cc")
        adv_label.pack(anchor="w", pady=(16, 0))
        self._method_adv_frame = ttk.LabelFrame(tab, text="Optimization", padding=8)
        adv_label.bind("<Button-1>", self._toggle_method_advanced)

        self._build_method_advanced(self._method_adv_frame)
        tab.columnconfigure(0, weight=1)
        return tab

    def _build_method_advanced(self, frame: ttk.Frame):
        ttk.Label(frame, text="Kriging n_trials:").grid(row=0, column=0, sticky="w")
        self._k_trials_var = tk.StringVar(value="300")
        ttk.Entry(frame, textvariable=self._k_trials_var, width=6).grid(row=0, column=1)
        self._k_trials_var.trace_add("write",
            lambda *_: self._try_int(self._k_trials_var, "kriging_n_trials"))

        ttk.Label(frame, text="Kriging CV splits:").grid(row=1, column=0, sticky="w", pady=2)
        self._k_splits_var = tk.StringVar(value="3")
        ttk.Entry(frame, textvariable=self._k_splits_var, width=6).grid(row=1, column=1)
        self._k_splits_var.trace_add("write",
            lambda *_: self._try_int(self._k_splits_var, "kriging_n_splits"))

        ttk.Label(frame, text="GP n_trials:").grid(row=2, column=0, sticky="w", pady=2)
        self._gp_trials_var = tk.StringVar(value="300")
        ttk.Entry(frame, textvariable=self._gp_trials_var, width=6).grid(row=2, column=1)
        self._gp_trials_var.trace_add("write",
            lambda *_: self._try_int(self._gp_trials_var, "gp_n_trials"))

    def _toggle_method_advanced(self, event=None):
        label = event.widget
        if self._method_adv_frame.winfo_ismapped():
            self._method_adv_frame.pack_forget()
            label.config(text="▸ Optimization settings")
        else:
            self._method_adv_frame.pack(fill="x", pady=(4, 0))
            label.config(text="▾ Optimization settings")

    # ─────────────────────────────────────────────────────────────────────────
    # Tab 3 — Model (switches Kriging ↔ GP panel)
    # ─────────────────────────────────────────────────────────────────────────

    def _build_tab_model(self) -> ttk.Frame:
        tab = ttk.Frame(self._notebook, padding=8)

        self._kriging_panel = KrigingPanel(tab, self.state)
        self._gp_panel      = GPPanel(tab, self.state)

        self._kriging_panel.pack(fill="both", expand=True)
        # GP panel hidden initially
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(0, weight=1)
        return tab

    def _switch_model_panel(self):
        mode = self.state["engine_mode"]
        if mode == "kriging":
            self._gp_panel.pack_forget()
            self._kriging_panel.pack(fill="both", expand=True)
        else:
            self._kriging_panel.pack_forget()
            self._gp_panel.pack(fill="both", expand=True)

    # ─────────────────────────────────────────────────────────────────────────
    # Tab 4 — Run & Results
    # ─────────────────────────────────────────────────────────────────────────

    def _build_tab_run(self) -> ttk.Frame:
        tab = ttk.Frame(self._notebook, padding=12)

        # ── Top row: Run / Cancel buttons + progress ──────────────────────────
        top = ttk.Frame(tab)
        top.pack(fill="x")

        self._run_btn = ttk.Button(top, text="▶  Run Interpolation",
                                   command=self._on_run, style="Accent.TButton")
        self._run_btn.pack(side="left", padx=4)

        self._cancel_btn = ttk.Button(top, text="■  Cancel",
                                       command=self._on_cancel, state="disabled")
        self._cancel_btn.pack(side="left", padx=4)

        self._progress = ttk.Progressbar(top, mode="indeterminate", length=220)
        self._progress.pack(side="left", padx=16)

        self._elapsed_var = tk.StringVar(value="")
        ttk.Label(top, textvariable=self._elapsed_var,
                  foreground="gray").pack(side="left")

        # ── Log window ────────────────────────────────────────────────────────
        ttk.Label(tab, text="Engine log:").pack(anchor="w", pady=(10, 2))
        log_frame = ttk.Frame(tab)
        log_frame.pack(fill="both", expand=True)

        self._log_text = tk.Text(log_frame, height=14, state="disabled",
                                  font=("Consolas", 9), wrap="word",
                                  background="#1e1e1e", foreground="#d4d4d4",
                                  insertbackground="white")
        vsb = ttk.Scrollbar(log_frame, orient="vertical",
                             command=self._log_text.yview)
        self._log_text.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._log_text.pack(side="left", fill="both", expand=True)

        # ── Results summary ───────────────────────────────────────────────────
        self._results_frame = ttk.LabelFrame(tab, text="Results", padding=8)
        self._results_frame.pack(fill="x", pady=(10, 0))

        self._summary_var = tk.StringVar(value="Run the engine to see results.")
        ttk.Label(self._results_frame, textvariable=self._summary_var,
                  justify="left").pack(anchor="w")

        btn_row = ttk.Frame(self._results_frame)
        btn_row.pack(anchor="w", pady=(6, 0))
        self._open_folder_btn = ttk.Button(btn_row, text="Open Results Folder",
                                            command=self._open_results_folder,
                                            state="disabled")
        self._open_folder_btn.pack(side="left", padx=(0, 8))

        # ── Collapsed CV table ────────────────────────────────────────────────
        cv_toggle = ttk.Label(self._results_frame, text="▸ Evaluation Metrics (per fold)",
                               cursor="hand2", foreground="#0063cc")
        cv_toggle.pack(anchor="w", pady=(6, 0))
        self._cv_frame = ttk.Frame(self._results_frame)
        cv_toggle.bind("<Button-1>", self._toggle_cv_table)

        tab.rowconfigure(1, weight=1)
        return tab

    def _toggle_cv_table(self, event=None):
        label = event.widget
        if self._cv_frame.winfo_ismapped():
            self._cv_frame.pack_forget()
            label.config(text="▸ Evaluation Metrics (per fold)")
        else:
            self._cv_frame.pack(fill="x", pady=(4, 0))
            label.config(text="▾ Evaluation Metrics (per fold)")

    # ─────────────────────────────────────────────────────────────────────────
    # Event handlers
    # ─────────────────────────────────────────────────────────────────────────

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Open input file",
            filetypes=[("Data files", "*.csv *.txt *.xlsx *.xls *.shp"),
                       ("All files", "*.*")])
        if path:
            self._input_path_var.set(path)
            self.state["input_filepath"] = path
            self._load_columns(path)

    def _browse_output(self):
        path = filedialog.askdirectory(title="Choose output folder",
                                        initialdir=self.state["output_dir"])
        if path:
            self._output_dir_var.set(path)
            self.state["output_dir"] = path

    def _browse_any(self, var: tk.StringVar, state_key: str):
        path = filedialog.askopenfilename(
            filetypes=[("Data files", "*.csv *.txt *.xlsx *.xls *.shp"),
                       ("All files", "*.*")])
        if path:
            var.set(path)
            self.state[state_key] = path

    def _load_columns(self, filepath: str):
        cols = _read_columns(filepath)
        for cb in (self._col_x_cb, self._col_y_cb, self._col_val_cb):
            cb.configure(values=cols)
        # Auto-guess common column names
        low = [c.lower() for c in cols]
        def _pick(hints):
            for h in hints:
                for i, c in enumerate(low):
                    if h in c:
                        return cols[i]
            return cols[0] if cols else ""
        self._col_x_var.set(_pick(["x", "lon", "east", "easting"]))
        self._col_y_var.set(_pick(["y", "lat", "north", "northing"]))
        self._col_val_var.set(_pick(["val", "z", "dep", "gwl", "value", "head"]))
        self._on_column_changed()

    def _on_column_changed(self):
        self.state["col_x"]     = self._col_x_var.get()
        self.state["col_y"]     = self._col_y_var.get()
        self.state["col_value"] = self._col_val_var.get()
        self._reload_variogram_data()

    def _reload_variogram_data(self):
        fp  = self.state.get("input_filepath", "")
        cx  = self.state.get("col_x", "")
        cy  = self.state.get("col_y", "")
        cz  = self.state.get("col_value", "")
        if not (fp and cx and cy and cz):
            return
        X, y, n = _load_xyz(fp, cx, cy, cz)
        if X is None:
            return
        self._kriging_panel.load_data(X, y)
        self._gp_panel.load_data(X, y)
        self._log(f"Loaded {n} points. Empirical variogram updated.")

    def _on_method_changed(self):
        self.state["engine_mode"] = self._method_var.get()
        self._switch_model_panel()

    def _sync_formats(self):
        self.state["export_formats"] = [
            fmt for fmt, v in self._fmt_vars.items() if v.get()
        ]

    def _sync_resolution(self, *_):
        try:
            self.state["resolution_m"] = float(self._res_var.get())
        except ValueError:
            pass

    def _sync_hull(self):
        try:
            self.state["convex_hull_buffer"] = float(self._hull_var.get())
        except ValueError:
            pass

    def _sync_minsep(self):
        val = self._minsep_var.get().strip()
        self.state["min_separation"] = float(val) if val else None

    def _try_int(self, var: tk.StringVar, state_key: str):
        try:
            self.state[state_key] = int(var.get())
        except ValueError:
            pass

    # ─────────────────────────────────────────────────────────────────────────
    # Auto Optimize
    # ─────────────────────────────────────────────────────────────────────────

    def _on_auto_optimize_requested(self, _event=None):
        if not self._validate_required_fields():
            return
        if self._opt_runner and self._opt_runner.is_alive():
            messagebox.showinfo("Busy", "Auto-optimization is already running.")
            return
        self._log("── Auto Optimize started (this runs Optuna; may take 1–5 min) ──")
        self._progress.start(12)
        self._opt_runner = AutoOptimizeRunner(self.state, self._log_queue)
        self._opt_runner.start()
        self.after(150, self._poll_auto_optimize)

    def _poll_auto_optimize(self):
        # Drain log queue
        try:
            while True:
                item = self._log_queue.get_nowait()
                if item is None:
                    break
                self._log(item)
        except queue.Empty:
            pass

        if self._opt_runner and self._opt_runner.is_alive():
            self.after(150, self._poll_auto_optimize)
            return

        self._progress.stop()
        if self._opt_runner and self._opt_runner.error:
            messagebox.showerror("Auto Optimize failed", self._opt_runner.error)
            return
        if self._opt_runner and self._opt_runner.params:
            params = self._opt_runner.params
            cv     = getattr(self._opt_runner, "cv_summary", None)
            mode   = self.state.get("engine_mode", "kriging")
            if mode == "kriging":
                self._kriging_panel.populate_from_params(params)
                if cv:
                    self._kriging_panel.set_cv_results(cv)
            else:
                self._gp_panel.populate_from_params(params)
                if cv:
                    self._gp_panel.set_cv_results(cv)
            self._log("── Auto Optimize complete — sliders updated ──")

    # ─────────────────────────────────────────────────────────────────────────
    # Run / Cancel
    # ─────────────────────────────────────────────────────────────────────────

    def _on_run(self):
        if not self._validate_required_fields():
            return
        if self._runner and self._runner.is_alive():
            messagebox.showinfo("Busy", "Engine is already running.")
            return

        # Clear log
        self._log_text.configure(state="normal")
        self._log_text.delete("1.0", "end")
        self._log_text.configure(state="disabled")
        self._summary_var.set("Running…")
        self._open_folder_btn.configure(state="disabled")

        self._run_btn.configure(state="disabled")
        self._cancel_btn.configure(state="normal")
        self._progress.start(12)

        self._runner = EngineRunner(self.state, self._log_queue)
        self._runner.start()
        self.after(100, self._poll_run)

    def _poll_run(self):
        # Drain log queue
        try:
            while True:
                item = self._log_queue.get_nowait()
                if item is None:
                    self._on_run_complete()
                    return
                self._log(item)
        except queue.Empty:
            pass

        self.after(100, self._poll_run)

    def _on_run_complete(self):
        self._progress.stop()
        self._run_btn.configure(state="normal")
        self._cancel_btn.configure(state="disabled")

        elapsed = self._runner.elapsed if self._runner else 0
        self._elapsed_var.set(f"{elapsed:.0f} s")

        if self._runner and self._runner.error:
            err_first_line = self._runner.error.split("\n")[0]
            self._summary_var.set(f"Error: {err_first_line}")
            messagebox.showerror("Engine error", self._runner.error)
            return

        result = self._runner.result if self._runner else {}
        self._display_results(result)

    def _on_cancel(self):
        if self._runner:
            self._runner.cancel()
        self._cancel_btn.configure(state="disabled")
        self._progress.stop()
        self._run_btn.configure(state="normal")
        self._log("── Run cancelled ──")

    # ─────────────────────────────────────────────────────────────────────────
    # Results display
    # ─────────────────────────────────────────────────────────────────────────

    def _display_results(self, result: dict):
        mode     = result.get("mode", "?")
        elapsed  = result.get("elapsed", 0)
        params   = result.get("params", {})
        rmse     = result.get("rmse")
        mae      = result.get("mae")
        r2       = result.get("r2")

        lines = [f"Mode: {mode.upper()}   Runtime: {elapsed:.0f} s"]
        if rmse is not None: lines.append(f"RMSE: {rmse:.4f}")
        if mae  is not None: lines.append(f"MAE:  {mae:.4f}")
        if r2   is not None: lines.append(f"R²:   {r2:.4f}")
        if params:
            bm = params.get("best_model") or params.get("kernel_type", "")
            if bm:
                lines.append(f"Best model / kernel: {bm}")
        self._summary_var.set("\n".join(lines))

        self._open_folder_btn.configure(state="normal")
        self._last_run_dir = result.get("run_dir", "")

        # Build CV table
        for w in self._cv_frame.winfo_children():
            w.destroy()
        cv_rows = result.get("cv_rows", [])
        if cv_rows:
            headers = list(cv_rows[0].keys())
            for j, h in enumerate(headers):
                ttk.Label(self._cv_frame, text=h, font=("TkDefaultFont", 9, "bold")).grid(
                    row=0, column=j, padx=4, pady=2, sticky="w")
            for i, row in enumerate(cv_rows, start=1):
                for j, h in enumerate(headers):
                    ttk.Label(self._cv_frame, text=row.get(h, ""),
                               font=("Consolas", 9)).grid(
                        row=i, column=j, padx=4, pady=1, sticky="w")

    def _open_results_folder(self):
        folder = getattr(self, "_last_run_dir", self.state.get("output_dir", ""))
        if folder and os.path.isdir(folder):
            if sys.platform == "win32":
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.run(["open", folder])
            else:
                subprocess.run(["xdg-open", folder])
        else:
            messagebox.showinfo("Not found", f"Folder does not exist:\n{folder}")

    # ─────────────────────────────────────────────────────────────────────────
    # Utilities
    # ─────────────────────────────────────────────────────────────────────────

    def _validate_required_fields(self) -> bool:
        if not self.state.get("input_filepath"):
            messagebox.showwarning("Input required", "Please select an input file first.")
            self._notebook.select(0)
            return False
        if not (self.state.get("col_x") and self.state.get("col_y")
                and self.state.get("col_value")):
            messagebox.showwarning("Columns required",
                                   "Please select X, Y, and Value columns.")
            self._notebook.select(0)
            return False
        return True

    def _log(self, text: str):
        self._log_text.configure(state="normal")
        self._log_text.insert("end", text + "\n")
        self._log_text.see("end")
        self._log_text.configure(state="disabled")


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
